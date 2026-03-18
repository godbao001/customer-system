# 业务管理路由
from permissions import check_permission, ALL_PERMISSIONS
from flask import Blueprint, render_template, request, jsonify, session, send_file
from models import db, Order, OrderItem, Product, ORDER_STATUS, FieldConfig, GenerateRecord
from datetime import datetime
from utils.log import add_log
from sqlalchemy import or_, and_
import os

biz_bp = Blueprint('biz', __name__, url_prefix='/biz')

# 生成制作页面
@biz_bp.route('/generate')
@check_permission('biz_generate')
def generate():
    """生成制作页面"""
    return render_template('biz/generate.html')

# 获取待制作订单列表（付款未制作）
@biz_bp.route('/api/pending_orders')
@check_permission('biz_generate')
def pending_orders():
    """获取所有付款未制作的订单"""
    orders = Order.query.filter_by(status=3, is_deleted=0).order_by(Order.id.desc()).all()
    
    # 获取所有经营模式配置（用field_name匹配）
    model_configs = {}
    configs = FieldConfig.query.filter_by(field_type='business_model', status=1).all()
    for config in configs:
        model_configs[config.field_name] = config.color or ''
    
    result = []
    for order in orders:
        order_dict = order.to_dict()
        # 获取店铺信息
        shop = order.shop
        business_model = ''
        model_color = ''
        
        if shop:
            order_dict['shop_display'] = f'[{shop.id}]{shop.shop_name}'
            business_model = shop.business_model or ''
            order_dict['shop_address'] = shop.address or ''
            order_dict['shop_phone'] = shop.phone or ''
            # 从配置中获取对应颜色
            model_color = model_configs.get(business_model, '')
        else:
            order_dict['shop_display'] = f'[{order.shop_id}]未知店铺'
            order_dict['shop_address'] = ''
            order_dict['shop_phone'] = ''
        
        order_dict['shop_business_model'] = business_model
        order_dict['shop_model_color'] = model_color
        result.append(order_dict)
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 获取待制作常规文宣（尺寸和备注没有变更）
@biz_bp.route('/api/regular_items')
@check_permission('biz_generate')
def regular_items():
    """获取所有付款未制作订单中的常规文宣（尺寸和备注无变更）- 合并同类产品"""
    # 先获取付款未制作的订单ID
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    
    # 特殊产品类型列表（这些不归类到常规文宣）
    special_types = ['传单', '门帘', '拱门', '收据发票', '其他特殊']
    
    # 查询订单明细：尺寸和备注无变更，排除特殊类型产品
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        or_(
            Product.product_type == None,
            Product.product_type.notin_(special_types)
        ),
        # 原始尺寸没有修改（宽=原始宽 AND 高=原始高）AND 备注为空
        OrderItem.width == OrderItem.original_width,
        OrderItem.height == OrderItem.original_height,
        or_(OrderItem.remark == None, OrderItem.remark == '')
    ).all()
    
    # 按产品名称+规格合并
    merged = {}
    for item in items:
        order = Order.query.get(item.order_id)
        if not order:
            continue
        
        shop = order.shop
        shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        
        # 合并key：产品名称 + 规格
        key = f"{item.product_name}|{item.width}|{item.height}"
        
        if key not in merged:
            merged[key] = {
                'product_name': item.product_name,
                'width': item.width,
                'height': item.height,
                'total_quantity': 0,
                'shops': []
            }
        
        merged[key]['total_quantity'] += item.quantity
        merged[key]['shops'].append(shop_name)
    
    # 转换为列表
    result = []
    for key, data in merged.items():
        result.append({
            'product_name': data['product_name'],
            'spec': f'{data["width"]} × {data["height"]}',
            'quantity': data['total_quantity'],
            'shops': '、'.join(data['shops'])
        })
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 获取待制作异型文宣（尺寸或备注有变更）
@biz_bp.route('/api/custom_items')
@check_permission('biz_generate')
def custom_items():
    """获取所有付款未制作订单中的异型文宣（尺寸或备注有变更）- 合并同类"""
    # 先获取付款未制作的订单ID
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    
    # 特殊产品类型列表（这些不归类到异型文宣）
    special_types = ['传单', '门帘', '拱门', '收据发票', '其他特殊']
    
    # 查询订单明细：尺寸修改过或备注有内容，排除特殊类型产品
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        or_(
            Product.product_type == None,
            Product.product_type.notin_(special_types)
        ),
        or_(
            # 原始尺寸修改过
            OrderItem.width != OrderItem.original_width,
            OrderItem.height != OrderItem.original_height,
            # 或者有备注内容
            and_(OrderItem.remark != None, OrderItem.remark != '')
        )
    ).all()
    
    # 按产品名称+现规格+备注合并
    merged = {}
    for item in items:
        order = Order.query.get(item.order_id)
        if not order:
            continue
        
        shop = order.shop
        shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        
        # 合并key：产品名称 + 现规格 + 备注
        remark = item.remark or ''
        key = f"{item.product_name}|{item.width}|{item.height}|{remark}"
        
        if key not in merged:
            merged[key] = {
                'product_name': item.product_name,
                'original_spec': f'{item.original_width} × {item.original_height}' if item.original_width > 0 and item.original_height > 0 else f'{item.width} × {item.height}',
                'current_spec': f'{item.width} × {item.height}',
                'remark': remark,
                'total_quantity': 0,
                'shops': []
            }
        
        merged[key]['total_quantity'] += item.quantity
        merged[key]['shops'].append(shop_name)
    
    # 转换为列表
    result = []
    for key, data in merged.items():
        result.append({
            'product_name': data['product_name'],
            'original_spec': data['original_spec'],
            'current_spec': data['current_spec'],
            'quantity': data['total_quantity'],
            'remark': data['remark'],
            'shops': '、'.join(data['shops'])
        })
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 获取待制作其他特殊产品
@biz_bp.route('/api/other_items')
@check_permission('biz_generate')
def other_items():
    """获取所有付款未制作订单中的其他特殊类型产品"""
    # 先获取付款未制作的订单ID
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    
    # 查询订单明细：产品类型为其他特殊
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        Product.product_type == '其他特殊'
    ).all()
    
    # 获取关联的订单信息
    result = []
    for item in items:
        order = Order.query.get(item.order_id)
        if order:
            shop = order.shop
            shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
            result.append({
                **item.to_dict(),
                'order_no': order.order_no,
                'shop_name': shop_name,
                'product_type': '其他特殊'
            })
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 获取待制作门帘产品
@biz_bp.route('/api/doorway_items')
@check_permission('biz_generate')
def doorway_items():
    """获取所有付款未制作订单中的门帘类型产品"""
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        Product.product_type == '门帘'
    ).all()
    
    result = []
    for item in items:
        order = Order.query.get(item.order_id)
        if order:
            shop = order.shop
            shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
            result.append({
                **item.to_dict(),
                'order_no': order.order_no,
                'shop_name': shop_name,
                'express_address': order.express_shop_address or '',
                'express_phone': order.express_shop_phone or '',
                'product_type': '门帘'
            })
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 获取待制作传单产品
@biz_bp.route('/api/flyer_items')
@check_permission('biz_generate')
def flyer_items():
    """获取所有付款未制作订单中的传单类型产品"""
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        Product.product_type == '传单'
    ).all()
    
    result = []
    for item in items:
        order = Order.query.get(item.order_id)
        if order:
            shop = order.shop
            shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
            result.append({
                **item.to_dict(),
                'order_no': order.order_no,
                'shop_name': shop_name,
                'express_address': order.express_shop_address or '',
                'express_phone': order.express_shop_phone or '',
                'product_type': '传单'
            })
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 获取待制作拱门产品
@biz_bp.route('/api/arch_items')
@check_permission('biz_generate')
def arch_items():
    """获取所有付款未制作订单中的拱门类型产品"""
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        Product.product_type == '拱门'
    ).all()
    
    result = []
    for item in items:
        order = Order.query.get(item.order_id)
        if order:
            shop = order.shop
            shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
            result.append({
                **item.to_dict(),
                'order_no': order.order_no,
                'shop_name': shop_name,
                'express_address': order.express_shop_address or '',
                'express_phone': order.express_shop_phone or '',
                'product_type': '拱门'
            })
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 获取待开发票收据产品
@biz_bp.route('/api/invoice_items')
@check_permission('biz_generate')
def invoice_items():
    """获取所有付款未制作订单中的收据发票类型产品"""
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        Product.product_type == '收据发票'
    ).all()
    
    result = []
    for item in items:
        order = Order.query.get(item.order_id)
        if order:
            shop = order.shop
            shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
            result.append({
                **item.to_dict(),
                'order_no': order.order_no,
                'shop_name': shop_name,
                'product_type': '收据发票'
            })
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': len(result),
        'data': result
    })

# 强制修改页面
@biz_bp.route('/force_edit')
@check_permission('biz_force_edit')
def force_edit():
    """强制修改页面"""
    return render_template('biz/force_edit.html')

# 获取订单状态列表
@biz_bp.route('/api/order_statuses')
@check_permission('biz_force_edit')
def order_statuses():
    """获取所有可用的订单状态"""
    return jsonify({
        'code': 0,
        'msg': 'success',
        'data': [{'id': k, 'name': v} for k, v in ORDER_STATUS.items()]
    })

# 根据ID获取订单详情
@biz_bp.route('/api/order/<int:order_id>')
@check_permission('biz_force_edit')
def get_order(order_id):
    """根据订单ID获取订单详情"""
    order = Order.query.filter_by(id=order_id, is_deleted=0).first()
    
    if not order:
        return jsonify({
            'code': 404,
            'msg': '订单不存在',
            'data': None
        })
    
    # 获取订单明细
    from models import OrderItem
    items = OrderItem.query.filter_by(order_id=order.id).all()
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'data': {
            **order.to_dict(),
            'items': [item.to_dict() for item in items]
        }
    })

# 强制修改订单状态
@biz_bp.route('/api/force_update_status', methods=['POST'])
@check_permission('biz_force_edit')
def force_update_status():
    """强制修改订单状态（无视业务规则）"""
    data = request.get_json()
    order_id = data.get('order_id')
    new_status = data.get('status')
    
    if not order_id:
        return jsonify({'code': 400, 'msg': '订单ID不能为空'})
    
    if new_status is None:
        return jsonify({'code': 400, 'msg': '状态不能为空'})
    
    order = Order.query.filter_by(id=order_id, is_deleted=0).first()
    if not order:
        return jsonify({'code': 404, 'msg': '订单不存在'})
    
    old_status = order.status
    old_status_text = ORDER_STATUS.get(old_status, '未知')
    new_status_text = ORDER_STATUS.get(new_status, '未知')
    
    # 直接修改状态，不做任何业务规则检查
    order.status = new_status
    order.updated_at = datetime.now()
    
    # 记录操作日志
    add_log(
        'order',
        f'强制修改订单状态: {order.order_no}',
        f'{old_status_text}({old_status}) → {new_status_text}({new_status})',
        'success'
    )
    
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'msg': f'订单状态已强制修改为: {new_status_text}',
        'data': {
            'old_status': old_status,
            'new_status': new_status,
            'old_status_text': old_status_text,
            'new_status_text': new_status_text
        }
    })

# 导出Excel
@biz_bp.route('/api/export_excel')
@check_permission('biz_generate')
def export_excel():
    """导出生成制作数据到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import base64
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_sheet(ws, header_row):
        """美化表格样式"""
        # 设置行高
        ws.row_dimensions[1].height = 25  # 表头行高
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        # 设置列宽自动调整
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大50字符
            ws.column_dimensions[column].width = adjusted_width
        
        # 表头样式
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 数据样式
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
    
    # 获取所有数据
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    special_types = ['传单', '门帘', '拱门', '收据发票', '其他特殊']
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # 1. 待制作订单
    ws1 = wb.create_sheet('待制作订单')
    orders = Order.query.filter(Order.status == 3, Order.is_deleted == 0).order_by(Order.id.desc()).all()
    
    header1 = ['ID', '店铺', '经营模式', '金额', '地址', '电话', '到付/寄付', '快递方式', '状态']
    ws1.append(header1)
    
    for order in orders:
        shop = order.shop
        shop_display = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        business_model = shop.business_model if shop else ''
        free_shipping_text = '包邮' if order.free_shipping == 1 else ('到付' if order.express_fee_type == '到付' else '寄付')
        ws1.append([
            order.id, shop_display, business_model, float(order.total_amount) if order.total_amount else 0,
            order.express_shop_address or '', order.express_shop_phone or '',
            free_shipping_text, order.express_method or '', order.get_status_text()
        ])
    style_sheet(ws1, 1)
    
    # 2. 待制作常规文宣
    ws2 = wb.create_sheet('待制作常规文宣')
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        or_(Product.product_type == None, Product.product_type.notin_(special_types)),
        OrderItem.width == OrderItem.original_width,
        OrderItem.height == OrderItem.original_height,
        or_(OrderItem.remark == None, OrderItem.remark == '')
    ).all()
    
    merged = {}
    for item in items:
        order = Order.query.get(item.order_id)
        if not order:
            continue
        shop = order.shop
        shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        key = f"{item.product_name}|{item.width}|{item.height}"
        if key not in merged:
            merged[key] = {'product_name': item.product_name, 'spec': f'{item.width}×{item.height}', 'quantity': 0, 'shops': []}
        merged[key]['quantity'] += item.quantity
        merged[key]['shops'].append(shop_name)
    
    header2 = ['产品名称', '规格(宽×高)', '数量', '订购店铺']
    ws2.append(header2)
    
    for key, data in merged.items():
        ws2.append([data['product_name'], data['spec'], data['quantity'], '、'.join(data['shops'])])
    style_sheet(ws2, 1)
    
    # 3. 待制作异型文宣
    ws3 = wb.create_sheet('待制作异型文宣')
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        or_(Product.product_type == None, Product.product_type.notin_(special_types)),
        or_(OrderItem.width != OrderItem.original_width, OrderItem.height != OrderItem.original_height)
    ).all()
    
    merged = {}
    for item in items:
        order = Order.query.get(item.order_id)
        if not order:
            continue
        shop = order.shop
        shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        remark = item.remark or ''
        key = f"{item.product_name}|{item.width}|{item.height}|{remark}"
        if key not in merged:
            merged[key] = {'product_name': item.product_name, 'original_spec': f'{item.original_width}×{item.original_height}',
                         'current_spec': f'{item.width}×{item.height}', 'quantity': 0, 'remark': remark, 'shops': []}
        merged[key]['quantity'] += item.quantity
        merged[key]['shops'].append(shop_name)
    
    header3 = ['产品名称', '原规格', '现规格', '数量', '备注', '订购店铺']
    ws3.append(header3)
    
    for key, data in merged.items():
        ws3.append([data['product_name'], data['original_spec'], data['current_spec'], data['quantity'], data['remark'], '、'.join(data['shops'])])
    style_sheet(ws3, 1)
    
    # 4-7. 特殊类型产品
    type_sheets = {
        '门帘': 'doorway_items',
        '传单': 'flyer_items', 
        '拱门': 'arch_items',
        '收据发票': 'invoice_items',
        '其他特殊': 'other_items'
    }
    
    type_headers = {
        '门帘': ['店铺', '产品名称', '规格(宽×高)', '数量', '地址', '电话', '备注'],
        '传单': ['店铺', '产品名称', '规格(宽×高)', '数量', '备注'],
        '拱门': ['店铺', '产品名称', '规格(宽×高)', '数量', '备注'],
        '收据发票': ['店铺', '产品名称', '规格(宽×高)', '数量', '小计'],
        '其他特殊': ['订单号', '店铺', '产品名称', '规格(宽×高)', '数量', '小计']
    }
    
    for type_name, api_name in type_sheets.items():
        ws = wb.create_sheet(f'待制作{type_name}')
        
        # 获取数据
        items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
            OrderItem.order_id.in_(order_ids),
            Product.product_type == type_name
        ).all()
        
        header = type_headers.get(type_name, ['店铺', '产品名称', '规格', '数量'])
        ws.append(header)
        
        for item in items:
            order = Order.query.get(item.order_id)
            if not order:
                continue
            shop = order.shop
            shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
            
            if type_name == '门帘':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity,
                          order.express_shop_address or '', order.express_shop_phone or '', item.remark or ''])
            elif type_name == '传单':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, item.remark or ''])
            elif type_name == '拱门':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, item.remark or ''])
            elif type_name == '收据发票':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, float(item.subtotal) if item.subtotal else 0])
            else:  # 其他特殊
                ws.append([order.order_no, shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, float(item.subtotal) if item.subtotal else 0])
        
        style_sheet(ws, 1)
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 返回base64
    b64 = base64.b64encode(output.getvalue()).decode()
    
    return jsonify({
        'code': 0,
        'msg': '导出成功',
        'data': f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}'
    })

# 生成制作文件并更新订单状态
@biz_bp.route('/api/generate_and_export')
@check_permission('biz_generate')
def generate_and_export():
    """生成制作文件并更新订单状态"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    import base64
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_sheet(ws, header_row):
        """美化表格样式"""
        ws.row_dimensions[1].height = 25
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
        
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
    
    # 获取所有数据
    order_ids = db.session.query(Order.id).filter(Order.status == 3, Order.is_deleted == 0).subquery()
    special_types = ['传单', '门帘', '拱门', '收据发票', '其他特殊']
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # 1. 待制作订单
    ws1 = wb.create_sheet('待制作订单')
    orders = Order.query.filter(Order.status == 3, Order.is_deleted == 0).order_by(Order.id.desc()).all()
    
    header1 = ['ID', '店铺', '经营模式', '金额', '地址', '电话', '到付/寄付', '快递方式', '状态']
    ws1.append(header1)
    
    for order in orders:
        shop = order.shop
        shop_display = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        business_model = shop.business_model if shop else ''
        free_shipping_text = '包邮' if order.free_shipping == 1 else ('到付' if order.express_fee_type == '到付' else '寄付')
        ws1.append([
            order.id, shop_display, business_model, float(order.total_amount) if order.total_amount else 0,
            order.express_shop_address or '', order.express_shop_phone or '',
            free_shipping_text, order.express_method or '', order.get_status_text()
        ])
    style_sheet(ws1, 1)
    
    # 2. 待制作常规文宣
    ws2 = wb.create_sheet('待制作常规文宣')
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        or_(Product.product_type == None, Product.product_type.notin_(special_types)),
        OrderItem.width == OrderItem.original_width,
        OrderItem.height == OrderItem.original_height,
        or_(OrderItem.remark == None, OrderItem.remark == '')
    ).all()
    
    merged = {}
    for item in items:
        order = Order.query.get(item.order_id)
        if not order:
            continue
        shop = order.shop
        shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        key = f"{item.product_name}|{item.width}|{item.height}"
        if key not in merged:
            merged[key] = {'product_name': item.product_name, 'spec': f'{item.width}×{item.height}', 'quantity': 0, 'shops': []}
        merged[key]['quantity'] += item.quantity
        merged[key]['shops'].append(shop_name)
    
    header2 = ['产品名称', '规格(宽×高)', '数量', '订购店铺']
    ws2.append(header2)
    
    for key, data in merged.items():
        ws2.append([data['product_name'], data['spec'], data['quantity'], '、'.join(data['shops'])])
    style_sheet(ws2, 1)
    
    # 3. 待制作异型文宣
    ws3 = wb.create_sheet('待制作异型文宣')
    items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
        OrderItem.order_id.in_(order_ids),
        or_(Product.product_type == None, Product.product_type.notin_(special_types)),
        or_(OrderItem.width != OrderItem.original_width, OrderItem.height != OrderItem.original_height)
    ).all()
    
    merged = {}
    for item in items:
        order = Order.query.get(item.order_id)
        if not order:
            continue
        shop = order.shop
        shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
        remark = item.remark or ''
        key = f"{item.product_name}|{item.width}|{item.height}|{remark}"
        if key not in merged:
            merged[key] = {'product_name': item.product_name, 'original_spec': f'{item.original_width}×{item.original_height}',
                         'current_spec': f'{item.width}×{item.height}', 'quantity': 0, 'remark': remark, 'shops': []}
        merged[key]['quantity'] += item.quantity
        merged[key]['shops'].append(shop_name)
    
    header3 = ['产品名称', '原规格', '现规格', '数量', '备注', '订购店铺']
    ws3.append(header3)
    
    for key, data in merged.items():
        ws3.append([data['product_name'], data['original_spec'], data['current_spec'], data['quantity'], data['remark'], '、'.join(data['shops'])])
    style_sheet(ws3, 1)
    
    # 4-7. 特殊类型产品
    type_sheets = {
        '门帘': 'doorway_items',
        '传单': 'flyer_items', 
        '拱门': 'arch_items',
        '收据发票': 'invoice_items',
        '其他特殊': 'other_items'
    }
    
    type_headers = {
        '门帘': ['店铺', '产品名称', '规格(宽×高)', '数量', '地址', '电话', '备注'],
        '传单': ['店铺', '产品名称', '规格(宽×高)', '数量', '备注'],
        '拱门': ['店铺', '产品名称', '规格(宽×高)', '数量', '备注'],
        '收据发票': ['店铺', '产品名称', '规格(宽×高)', '数量', '小计'],
        '其他特殊': ['订单号', '店铺', '产品名称', '规格(宽×高)', '数量', '小计']
    }
    
    for type_name, api_name in type_sheets.items():
        ws = wb.create_sheet(f'待制作{type_name}')
        
        items = OrderItem.query.join(Product, OrderItem.product_id == Product.id).filter(
            OrderItem.order_id.in_(order_ids),
            Product.product_type == type_name
        ).all()
        
        header = type_headers.get(type_name, ['店铺', '产品名称', '规格', '数量'])
        ws.append(header)
        
        for item in items:
            order = Order.query.get(item.order_id)
            if not order:
                continue
            shop = order.shop
            shop_name = f'[{shop.id}]{shop.shop_name}' if shop else f'[{order.shop_id}]未知'
            
            if type_name == '门帘':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity,
                          order.express_shop_address or '', order.express_shop_phone or '', item.remark or ''])
            elif type_name == '传单':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, item.remark or ''])
            elif type_name == '拱门':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, item.remark or ''])
            elif type_name == '收据发票':
                ws.append([shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, float(item.subtotal) if item.subtotal else 0])
            else:
                ws.append([order.order_no, shop_name, item.product_name or '', f'{item.width}×{item.height}', item.quantity, float(item.subtotal) if item.subtotal else 0])
        
        style_sheet(ws, 1)
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # === 更新订单状态：付款未制作(3) -> 制作未打包(4) ===
    # 先获取订单总金额
    orders_to_update = Order.query.filter(Order.status == 3, Order.is_deleted == 0).all()
    total_amount = sum(float(o.total_amount or 0) for o in orders_to_update)
    updated_count = Order.query.filter(Order.status == 3, Order.is_deleted == 0).update({'status': 4, 'make_time': datetime.now()})
    db.session.commit()
    
    # 保存文件到服务器
    upload_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'uploads', 'generate')
    os.makedirs(upload_dir, exist_ok=True)
    
    # 生成文件名
    now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'制作文件_{now_str}.xlsx'
    file_path = os.path.join(upload_dir, filename)
    
    # 保存文件
    output.seek(0)
    with open(file_path, 'wb') as f:
        f.write(output.getvalue())
    
    # 保存记录
    operator = session.get('username', '未知用户')
    record = GenerateRecord(
        filename=filename,
        file_path=f'/static/uploads/generate/{filename}',
        order_count=updated_count,
        total_amount=total_amount,
        created_by=operator
    )
    db.session.add(record)
    db.session.commit()
    
    # 记录日志
    add_log('biz', '生成制作文件', f'生成制作Excel并更新{updated_count}个订单状态: 付款未制作→制作未打包', 'success')
    
    # 返回base64
    b64 = base64.b64encode(output.getvalue()).decode()
    
    return jsonify({
        'code': 0,
        'msg': f'生成成功！已更新{updated_count}个订单状态',
        'data': f'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}'
    })

# 生成制作记录页面
@biz_bp.route('/records')
@check_permission('biz_generate')
def records():
    """生成制作记录页面"""
    return render_template('biz/records.html')

# 获取生成记录列表
@biz_bp.route('/api/records')
@check_permission('biz_generate')
def api_records():
    """获取生成记录列表"""
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 20, type=int)
    
    query = GenerateRecord.query.order_by(GenerateRecord.id.desc())
    pagination = query.paginate(page=page, per_page=limit, error_out=False)
    
    return jsonify({
        'code': 0,
        'msg': 'success',
        'count': pagination.total,
        'data': [r.to_dict() for r in pagination.items]
    })

# 下载生成记录文件
@biz_bp.route('/api/records/<int:record_id>/download')
@check_permission('biz_generate')
def download_record(record_id):
    """下载生成记录文件"""
    record = GenerateRecord.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'})
    
    # 拼完整路径
    base_dir = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(base_dir, 'static', record.file_path.lstrip('/static/'))
    if not os.path.exists(file_path):
        return jsonify({'code': 404, 'msg': '文件不存在'})
    
    return send_file(file_path, as_attachment=True, download_name=record.filename)

# 删除生成记录
@biz_bp.route('/api/records/<int:record_id>', methods=['DELETE'])
@check_permission('biz_generate')
def delete_record(record_id):
    """删除生成记录"""
    record = GenerateRecord.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'})
    
    # 删除文件
    base_dir = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(base_dir, 'static', record.file_path.lstrip('/static/'))
    if os.path.exists(file_path):
        os.remove(file_path)
    
    # 删除记录
    db.session.delete(record)
    db.session.commit()
    
    # 记录日志
    add_log('biz', '删除生成记录', f'删除生成记录: {record.filename}', 'success')
    
    return jsonify({'code': 0, 'msg': '删除成功'})

# 订单打包页面
@biz_bp.route('/package')
@check_permission('biz_package')
def package():
    """订单打包页面"""
    return render_template('biz/package.html')
